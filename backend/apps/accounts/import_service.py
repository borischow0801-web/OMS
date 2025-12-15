"""
用户批量导入服务
"""
import logging
from typing import Dict, List, Tuple, Optional
from django.contrib.auth import get_user_model
from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

User = get_user_model()
logger = logging.getLogger(__name__)


class UserImportService:
    """用户批量导入服务类"""
    
    # 角色映射：Excel中的角色名称 -> 数据库中的角色代码
    ROLE_MAPPING = {
        '使用方': 'user',
        'user': 'user',
        '管理方': 'admin',
        'admin': 'admin',
        '承建方-项目经理': 'manager',
        'manager': 'manager',
        '项目经理': 'manager',
        '承建方-员工': 'employee',
        'employee': 'employee',
        '员工': 'employee',
    }
    
    # Excel列名映射
    COLUMN_MAPPING = {
        '用户名': 'username',
        'username': 'username',
        '密码': 'password',
        'password': 'password',
        '姓名': '姓名',  # 特殊处理，单独存储
        'first_name': 'first_name',
        '名': 'first_name',
        'last_name': 'last_name',
        '姓': 'last_name',
        '邮箱': 'email',
        'email': 'email',
        '角色': 'role',
        'role': 'role',
        '手机号': 'phone',
        'phone': 'phone',
        '部门': 'department',
        'department': 'department',
    }
    
    @classmethod
    def parse_excel(cls, file_path: str) -> Tuple[List[Dict], List[str]]:
        """
        解析Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Tuple[List[Dict], List[str]]: (用户数据列表, 错误信息列表)
        """
        errors = []
        users_data = []
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # 读取表头
            headers = []
            header_row = 1
            for cell in sheet[header_row]:
                header_value = str(cell.value).strip() if cell.value else ''
                headers.append(header_value)
            
            # 检查必需列
            required_columns = ['用户名', '密码']
            missing_columns = []
            column_indices = {}
            
            for idx, header in enumerate(headers, start=1):
                normalized_header = header.strip()
                if normalized_header in cls.COLUMN_MAPPING:
                    column_indices[cls.COLUMN_MAPPING[normalized_header]] = idx
                elif normalized_header in cls.COLUMN_MAPPING.values():
                    column_indices[normalized_header] = idx
            
            # 检查必需列是否存在
            if 'username' not in column_indices:
                errors.append(f'缺少必需列：用户名')
                return users_data, errors
            if 'password' not in column_indices:
                errors.append(f'缺少必需列：密码')
                return users_data, errors
            
            # 读取数据行
            row_num = 2
            while True:
                row_data = {}
                has_data = False
                
                for field, col_idx in column_indices.items():
                    cell_value = sheet.cell(row=row_num, column=col_idx).value
                    if cell_value is not None:
                        has_data = True
                        row_data[field] = str(cell_value).strip() if cell_value else ''
                    else:
                        row_data[field] = ''
                
                if not has_data:
                    break
                
                # 跳过空行
                if not row_data.get('username'):
                    row_num += 1
                    continue
                
                users_data.append({
                    'row_num': row_num,
                    'data': row_data
                })
                row_num += 1
            
            workbook.close()
            
        except Exception as e:
            errors.append(f'解析Excel文件失败：{str(e)}')
            logger.error(f'解析Excel文件失败：{e}', exc_info=True)
        
        return users_data, errors
    
    @classmethod
    def validate_user_data(cls, user_data: Dict, row_num: int) -> List[str]:
        """
        验证单条用户数据
        
        Args:
            user_data: 用户数据字典
            row_num: 行号（用于错误提示）
            
        Returns:
            List[str]: 错误信息列表
        """
        errors = []
        
        # 验证用户名
        username = user_data.get('username', '').strip()
        if not username:
            errors.append(f'第{row_num}行：用户名为空')
        elif len(username) > 150:
            errors.append(f'第{row_num}行：用户名长度不能超过150个字符')
        elif User.objects.filter(username=username).exists():
            errors.append(f'第{row_num}行：用户名"{username}"已存在')
        
        # 验证密码
        password = user_data.get('password', '').strip()
        if not password:
            errors.append(f'第{row_num}行：密码为空')
        elif len(password) < 8:
            errors.append(f'第{row_num}行：密码长度至少8个字符')
        
        # 验证邮箱格式（如果提供）
        email = user_data.get('email', '').strip()
        if email:
            from django.core.validators import validate_email
            try:
                validate_email(email)
                # 邮箱格式正确，检查是否已被使用
                if User.objects.filter(email=email).exists():
                    errors.append(f'第{row_num}行：邮箱"{email}"已被使用')
            except ValidationError:
                errors.append(f'第{row_num}行：邮箱格式不正确')
        
        # 验证角色
        role_str = user_data.get('role', '').strip()
        if role_str:
            if role_str not in cls.ROLE_MAPPING:
                valid_roles = ', '.join(set(cls.ROLE_MAPPING.keys()) - {'user', 'admin', 'manager', 'employee'})
                errors.append(f'第{row_num}行：角色"{role_str}"无效，有效值为：{valid_roles}')
        
        return errors
    
    @classmethod
    def normalize_role(cls, role_str: str) -> str:
        """规范化角色代码"""
        role_str = role_str.strip()
        return cls.ROLE_MAPPING.get(role_str, 'user')
    
    @classmethod
    @transaction.atomic
    def import_users(cls, file_path: str) -> Dict[str, any]:
        """
        批量导入用户
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict: 导入结果，包含成功数量、失败数量和错误信息
        """
        result = {
            'success_count': 0,
            'failed_count': 0,
            'errors': [],
            'details': []
        }
        
        # 解析Excel文件
        users_data, parse_errors = cls.parse_excel(file_path)
        result['errors'].extend(parse_errors)
        
        if parse_errors:
            result['failed_count'] = len(users_data)
            return result
        
        # 验证并导入用户
        all_validation_errors = []
        valid_users = []
        
        for item in users_data:
            row_num = item['row_num']
            user_data = item['data']
            
            # 验证数据
            validation_errors = cls.validate_user_data(user_data, row_num)
            
            if validation_errors:
                all_validation_errors.extend(validation_errors)
                result['details'].append({
                    'row_num': row_num,
                    'username': user_data.get('username', ''),
                    'status': '失败',
                    'errors': validation_errors
                })
            else:
                valid_users.append({
                    'row_num': row_num,
                    'data': user_data
                })
        
        # 如果有验证错误，不进行导入
        if all_validation_errors:
            result['errors'].extend(all_validation_errors)
            result['failed_count'] = len(users_data)
            return result
        
        # 导入用户
        for item in valid_users:
            row_num = item['row_num']
            user_data = item['data']
            username = user_data.get('username', '').strip()
            
            try:
                # 处理姓名字段
                first_name = user_data.get('first_name', '').strip() or ''
                last_name = user_data.get('last_name', '').strip() or ''
                
                # 如果只有"姓名"字段，将其作为first_name
                if not first_name and not last_name:
                    full_name = user_data.get('姓名', '').strip()
                    if full_name:
                        # 简单处理：将姓名作为first_name
                        first_name = full_name
                
                # 创建用户
                user = User(
                    username=username,
                    email=user_data.get('email', '').strip() or '',
                    first_name=first_name,
                    last_name=last_name,
                    phone=user_data.get('phone', '').strip() or None,
                    department=user_data.get('department', '').strip() or None,
                    role=cls.normalize_role(user_data.get('role', 'user')),
                    is_active=True
                )
                
                # 设置密码
                password = user_data.get('password', '').strip()
                user.set_password(password)
                
                # 验证并保存
                user.full_clean()
                user.save()
                
                result['success_count'] += 1
                result['details'].append({
                    'row_num': row_num,
                    'username': username,
                    'status': '成功'
                })
                
            except Exception as e:
                error_msg = f'第{row_num}行：创建用户失败 - {str(e)}'
                result['errors'].append(error_msg)
                result['failed_count'] += 1
                result['details'].append({
                    'row_num': row_num,
                    'username': username,
                    'status': '失败',
                    'errors': [error_msg]
                })
                logger.error(f'导入用户失败（第{row_num}行）：{e}', exc_info=True)
        
        return result

