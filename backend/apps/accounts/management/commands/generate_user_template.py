"""
Django管理命令：生成用户导入Excel模板文件

使用方法：
    python manage.py generate_user_template
"""
from django.core.management.base import BaseCommand
from django.conf import settings
import os


class Command(BaseCommand):
    help = '生成用户批量导入Excel模板文件'

    def handle(self, *args, **options):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            self.stdout.write(
                self.style.ERROR('错误：请先安装openpyxl库')
            )
            self.stdout.write('运行命令：pip install openpyxl==3.1.2')
            return
        
        # 创建Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "用户导入模板"
        
        # 表头
        headers = [
            ('用户名', '必填，唯一，不能与现有用户重复'),
            ('密码', '必填，至少8个字符'),
            ('姓名', '可选，用户显示名称'),
            ('邮箱', '可选，唯一，不能与现有用户重复'),
            ('角色', '可选，使用方/管理方/承建方-项目经理/承建方-员工，默认为使用方'),
            ('手机号', '可选'),
            ('部门', '可选'),
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 说明行样式
        note_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        note_font = Font(size=9, color="333333")
        note_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 写入表头
        for col_idx, (header, note) in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            
            # 表头单元格
            header_cell = ws[f'{col_letter}1']
            header_cell.value = header
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment
            
            # 说明单元格
            note_cell = ws[f'{col_letter}2']
            note_cell.value = note
            note_cell.fill = note_fill
            note_cell.font = note_font
            note_cell.alignment = note_alignment
        
        # 添加示例数据行
        example_data = [
            ('zhangsan', '12345678', '张三', 'zhangsan@example.com', '使用方', '13800138000', '信息中心'),
            ('lisi', '12345678', '李四', 'lisi@example.com', '管理方', '13800138001', '办公室'),
            ('wangwu', '12345678', '王五', 'wangwu@example.com', '承建方-项目经理', '13800138002', '技术部'),
            ('zhaoliu', '12345678', '赵六', 'zhaoliu@example.com', '承建方-员工', '13800138003', '技术部'),
        ]
        
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        example_alignment = Alignment(horizontal="left", vertical="center")
        
        for row_idx, row_data in enumerate(example_data, start=3):
            for col_idx, cell_value in enumerate(row_data, start=1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f'{col_letter}{row_idx}']
                cell.value = cell_value
                cell.fill = example_fill
                cell.alignment = example_alignment
        
        # 设置列宽
        column_widths = {
            'A': 15,  # 用户名
            'B': 15,  # 密码
            'C': 12,  # 姓名
            'D': 25,  # 邮箱
            'E': 20,  # 角色
            'F': 15,  # 手机号
            'G': 15,  # 部门
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 设置行高
        ws.row_dimensions[1].height = 25  # 表头行
        ws.row_dimensions[2].height = 40  # 说明行
        
        # 添加数据验证（角色下拉选项）
        role_options = ['使用方', '管理方', '承建方-项目经理', '承建方-员工']
        role_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(role_options)}"',
            allow_blank=True
        )
        role_dv.error = '请从下拉列表中选择角色'
        role_dv.errorTitle = '无效的角色'
        ws.add_data_validation(role_dv)
        role_dv.add(f'E3:E1000')  # 应用到E列的第3行到1000行
        
        # 保存文件到docs目录
        BASE_DIR = settings.BASE_DIR.parent  # OMS根目录
        template_dir = os.path.join(BASE_DIR, 'docs')
        os.makedirs(template_dir, exist_ok=True)
        
        template_path = os.path.join(template_dir, '用户导入模板.xlsx')
        wb.save(template_path)
        
        self.stdout.write(
            self.style.SUCCESS(f'✓ 模板文件已生成：{template_path}')
        )
        self.stdout.write('您可以在后台管理界面使用此模板进行用户批量导入。')

